# timesheetbot_agent/generators/govtech_excel.py
from __future__ import annotations

from datetime import datetime, timedelta
from calendar import monthrange
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border

# Try relative PH import first (package usage), then absolute (direct script use)
try:
    from ..ph_sg import PUBLIC_HOLIDAYS as PH_DEFAULT  # type: ignore
except Exception:  # pragma: no cover
    try:
        from ph_sg import PUBLIC_HOLIDAYS as PH_DEFAULT  # type: ignore
    except Exception:  # pragma: no cover
        PH_DEFAULT: Dict[str, str] = {}

# Reuse your exact styles
from ..styles import (
    thin_border, white_fill, yellow_fill, light_green_fill, lighter_green_fill,
    light_yellow_fill, light_blue_fill, light_red_fill, bold_font, red_font,
    black_font, center_alignment, right_alignment
)

# ---------- Helpers ----------

def _expand_leaves(
    leave_details: List[Sequence],
    year: int
) -> List[Tuple[str, str]]:
    """
    Accepts:
      - (start '%d-%B', end '%d-%B' or None, leave_type)
      - (date '%d-%B', leave_type)
    where the container can be a tuple **or** list.
    Returns list of (YYYY-MM-DD, leave_type)
    """
    expanded: List[Tuple[str, str]] = []

    for entry in leave_details:
        if isinstance(entry, (list, tuple)) and len(entry) == 3:
            start_s, end_s, leave_type = entry
            if not end_s:
                end_s = start_s

            start_dt = datetime.strptime(str(start_s), "%d-%B").replace(year=year)
            end_dt = datetime.strptime(str(end_s), "%d-%B").replace(year=year)

            while start_dt <= end_dt:
                expanded.append((start_dt.strftime("%Y-%m-%d"), str(leave_type)))
                start_dt += timedelta(days=1)

        elif isinstance(entry, (list, tuple)) and len(entry) == 2:
            date_s, leave_type = entry
            dt = datetime.strptime(str(date_s), "%d-%B").replace(year=year)
            expanded.append((dt.strftime("%Y-%m-%d"), str(leave_type)))

        # Ignore malformed entries silently

    return expanded


def _col_letter(idx: int) -> str:
    """1-based column index -> Excel column letter."""
    div, mod = divmod(idx - 1, 26)
    letter = chr(65 + mod)
    if div == 0:
        return letter
    return _col_letter(div) + letter


# ---------- Main generator ----------

def generate_govtech_timesheet(
    profile: Dict,                   # registration data
    month: int,                      # 1..12
    year: int,                       # e.g., 2025
    leave_details: List[Sequence],   # [(start,end,type)] or [(date,type)] — list or tuple
    out_dir: str | Path = "generated_timesheets",
    public_holidays: Optional[Dict[str, str]] = None,  # {"YYYY-MM-DD": "Holiday Name"}
    remarks: Optional[Dict[str, str]] = None,          # {"DD-Month": "text"} or {"YYYY-MM-DD": "text"}
) -> str:
    """
    Builds a GovTech Excel timesheet mirroring your legacy format.
    Returns absolute path to the saved .xlsx file.
    """

    # Required profile fields
    name = profile["name"]
    skill_level = profile["skill_level"]
    role_specialization = profile["role_specialization"]
    group_specialization = profile["group_specialization"]
    contractor = profile["contractor"]
    po_ref = profile["po_ref"]
    po_date = profile["po_date"]
    description = profile["description"]
    reporting_officer = profile["reporting_officer"]
    timesheet_preference = float(profile.get("timesheet_preference", 1.0))  # 8.5 or 1.0

    # Public holidays (fallback to packaged defaults)
    PH: Dict[str, str] = public_holidays if public_holidays is not None else PH_DEFAULT
    # Remarks map
    REMARKS: Dict[str, str] = remarks or {}

    # Whether to show NS column
    ns_leave_present = any(
        (isinstance(x, (list, tuple)) and len(x) == 3 and x[2] == "NS Leave") or
        (isinstance(x, (list, tuple)) and len(x) == 2 and x[1] == "NS Leave")
        for x in leave_details
    )

    # File setup
    month_name = datetime(year, month, 1).strftime("%B")
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{month_name}_{year}_Timesheet_{name.replace(' ', '_')}.xlsx"
    out_path = out_dir / filename

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year} Timesheet"

    # ---------- Header blocks ----------
    ws.merge_cells("B2:D2")  # Description value
    ws.merge_cells("B3:D3")  # PO Ref value
    ws.merge_cells("B4:D4")  # PO Date value
    ws.merge_cells("G2:H2")  # Month/Year value
    ws.merge_cells("G3:H3")  # Contractor value

    ws["A2"], ws["B2"] = "Description", description
    ws["A3"], ws["B3"] = "PO Ref", po_ref
    ws["A4"], ws["B4"] = "PO Date", po_date
    ws["F2"], ws["G2"] = "Month/Year", f"{month_name} - {year}"
    ws["F3"], ws["G3"] = "Contractor", contractor

    # Borders for A-D rows 2..4
    for r in range(2, 5):
        for c in ("A", "B", "C", "D"):
            ws[f"{c}{r}"].border = thin_border

    # Borders for F-H rows 2..3
    for r in (2, 3):
        for c in ("F", "G", "H"):
            ws[f"{c}{r}"].border = thin_border

    # Yellow fill for values (B-D) and Month/Year (G2:H2)
    for r in range(2, 5):
        for c in ("B", "C", "D"):
            ws[f"{c}{r}"].fill = yellow_fill
    for c in ("G", "H"):
        ws[f"{c}2"].fill = yellow_fill
        ws[f"{c}3"].fill = PatternFill(fill_type=None)  # Contractor not yellow

    # Ensure column E is blank (no fill, no border)
    for r in range(2, 5):
        ws[f"E{r}"].border = Border()
        ws[f"E{r}"].fill = PatternFill(fill_type=None)

    # Alignment for header values
    for r in range(2, 5):
        for c in ("B", "C", "D"):
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="bottom")
    for r in (2, 3):
        for c in ("G", "H"):
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="bottom")
    for r in range(2, 5):
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="bottom")

    # Column widths (match legacy + ensure headers aren’t truncated)
    ws.column_dimensions["A"].width = 20   # SN (wider)
    ws.column_dimensions["B"].width = 25   # Date / value blocks
    ws.column_dimensions["C"].width = 10   # At Work
    ws.column_dimensions["D"].width = 22   # Public Holiday (so header fits)
    ws.column_dimensions["E"].width = 14   # Sick Leave (wider per request)
    ws.column_dimensions["F"].width = 20   # Childcare Leave (wider per request)
    ws.column_dimensions["G"].width = 20   # Annual Leave
    ws.column_dimensions["H"].width = 16   # Remarks (base)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---------- User details ----------
    ws.merge_cells("B6:D6")  # Name
    ws.merge_cells("B7:D7")  # Role
    ws.merge_cells("B8:D8")  # Group
    ws.merge_cells("G6:H6")  # Skill Level

    ws["A6"], ws["B6"] = "Name", name
    ws["A7"], ws["B7"] = "Role Specialization", role_specialization
    ws["A8"], ws["B8"] = "Group/Specialization", group_specialization
    ws["F6"], ws["G6"] = "Skill Level", skill_level

    # Borders
    for r in range(6, 9):
        for c in ("A", "B", "C", "D"):
            ws[f"{c}{r}"].border = thin_border
    for c in ("F", "G", "H"):
        ws[f"{c}6"].border = thin_border

    # Yellow fills for values
    for r in range(6, 9):
        for c in ("B", "C", "D"):
            ws[f"{c}{r}"].fill = yellow_fill
    for c in ("G", "H"):
        ws[f"{c}6"].fill = yellow_fill

    # Keep E blank
    for r in range(6, 9):
        ws[f"E{r}"].border = Border()
        ws[f"E{r}"].fill = PatternFill(fill_type=None)

    # Left align merged B cells; center the Skill Level value
    for r in range(6, 9):
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="bottom")
    ws["G6"].alignment = center_alignment

    # ---------- Table headers ----------
    headers = ["SN", "Date", "At Work", "Public Holiday", "Sick Leave", "Childcare Leave", "Annual Leave"]
    fills   = [ white_fill, white_fill, light_green_fill, light_yellow_fill,
                lighter_green_fill, white_fill, light_blue_fill ]

    if ns_leave_present:
        headers.append("National Service Leave")
        fills.append(light_red_fill)

    headers.append("Remarks")
    fills.append(white_fill)

    # All headers non-bold to match your latest requirement
    NON_BOLD = set(headers)
    for col_idx, (hdr, fill) in enumerate(zip(headers, fills), start=1):
        cell = ws.cell(row=10, column=col_idx, value=hdr)
        cell.font = Font(name="Arial", size=12, bold=(hdr not in NON_BOLD), color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.fill = fill

    ws["A10"].font = Font(name="Arial", size=12, bold=False, color="000000")
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")

    # NS width tweak
    if ns_leave_present:
        ws.column_dimensions["I"].width = 15  # NS column width

    # ---------- Data rows ----------
    expanded = _expand_leaves(leave_details, year)
    _, days_in_month = monthrange(year, month)

    totals = {
        "At Work": 0.0,
        "Public Holiday": 0.0,
        "Sick Leave": 0.0,
        "Childcare Leave": 0.0,
        "Annual Leave": 0.0,
    }
    if ns_leave_present:
        totals["National Service Leave"] = 0.0

    start_row = 11
    sn = 1

    # Column indices
    at_work_col = 3
    ph_col = 4
    sick_col = 5
    cc_col = 6
    al_col = 7
    ns_col = 8 if ns_leave_present else None
    remarks_col = 9 if ns_leave_present else 8

    # Paint Date column (B) yellow (legacy behavior)
    for r in range(start_row, start_row + days_in_month):
        ws[f"B{r}"].fill = yellow_fill

    for day in range(1, days_in_month + 1):
        r = start_row + (day - 1)
        date_obj = datetime(year, month, day)
        ymd = date_obj.strftime("%Y-%m-%d")
        disp_date = date_obj.strftime("%d-%B-%Y")
        weekday = date_obj.weekday()  # Mon=0..Sun=6
        key_dd_mon = date_obj.strftime("%d-%B")  # e.g., "11-August"

        # Defaults
        if timesheet_preference == 8.5:
            at_work = 8.5 if weekday in (0, 1, 2, 3) else 8.0  # Fri=8.0
        else:
            at_work = timesheet_preference if weekday not in (5, 6) else 0.0

        ph = 0.0
        sick = 0.0
        cc = 0.0
        al = 0.0
        ns = 0.0
        remark = "-"

        # Weekends
        if weekday == 5:
            at_work = 0.0
            remark = "Saturday"
        elif weekday == 6:
            at_work = 0.0
            remark = "Sunday"

        # Public holiday
        if ymd in PH:
            at_work = 0.0
            ph = 1.0
            remark = PH[ymd]

        # Apply leaves
        for leave_date, ltype in expanded:
            if leave_date != ymd:
                continue

            if ltype == "Sick Leave":
                if weekday not in (5, 6) and ymd not in PH:
                    sick = 1.0
                    at_work = 0.0

            elif ltype == "Childcare Leave":
                if weekday not in (5, 6) and ymd not in PH:
                    cc = 1.0
                    at_work = 0.0

            elif ltype == "Annual Leave":
                if weekday not in (5, 6) and ymd not in PH:
                    al = 1.0
                    at_work = 0.0

            elif ltype == "NS Leave":
                if weekday not in (5, 6) and ymd not in PH:
                    ns = 1.0
                    at_work = 0.0

            elif ltype == "Weekend Efforts":
                if weekday in (5, 6) or ymd in PH:
                    at_work = 8.0 if timesheet_preference == 8.5 else 1.0

            elif ltype == "Public Holiday Efforts":
                if ymd in PH:
                    at_work = 8.0 if timesheet_preference == 8.5 else 1.0

            elif ltype == "Half Day":
                if timesheet_preference == 8.5:
                    at_work = 4.0 if weekday == 4 else 4.5
                else:
                    at_work = 0.5

        # ----- Merge custom remarks (user comments) -----
        # Priority: user comment > weekend/PH text.
        custom_remark = REMARKS.get(key_dd_mon) or REMARKS.get(ymd)
        if custom_remark:
            remark = custom_remark

        # Totals
        totals["At Work"] += at_work if isinstance(at_work, float) else 0.0
        totals["Public Holiday"] += ph
        totals["Sick Leave"] += sick
        totals["Childcare Leave"] += cc
        totals["Annual Leave"] += al
        if ns_leave_present:
            totals["National Service Leave"] += ns

        # Row values (blank zeros, PH column uses "-" for 0.0 like legacy)
        row_vals: List = [
            sn,
            disp_date,
            "" if at_work == 0.0 else at_work,
            "-" if ph == 0.0 else ph,
            "" if sick == 0.0 else sick,
            "" if cc == 0.0 else cc,
            "" if al == 0.0 else al,
        ]
        if ns_leave_present:
            row_vals.append("" if ns == 0.0 else ns)

        row_vals.append(remark)

        # Write row
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = center_alignment
            cell.border = thin_border

        # Yellow fill for leave columns (not PH)
        leave_cols = [at_work_col, sick_col, cc_col, al_col]
        if ns_leave_present:
            leave_cols.append(ns_col)

        for c_idx in leave_cols:
            ws.cell(row=r, column=c_idx).fill = yellow_fill

        # Ensure PH column not yellow
        ws.cell(row=r, column=ph_col).fill = PatternFill(fill_type=None)

        # Right align numeric & PH columns
        for c_idx in [at_work_col, ph_col, sick_col, cc_col, al_col] + ([ns_col] if ns_leave_present else []):
            ws.cell(row=r, column=c_idx).alignment = right_alignment

        # Remarks styling:
        # rem_cell = ws.cell(row=r, column=remarks_col)
        # rem_cell.alignment = right_alignment
        # if rem_cell.value not in ["", "-"]:
        #     # Weekend/PH -> red with light red background
        #     if (ymd in PH) or (remark in ("Saturday", "Sunday")):
        #         rem_cell.fill = light_red_fill
        #         rem_cell.font = Font(name="Arial", size=12, color="FF0000", bold=False)
        #     else:
        #         # User comment or other – no red background, normal black text
        #         rem_cell.font = Font(name="Arial", size=12, color="000000", bold=False)
        # else:
        #     rem_cell.font = black_font
        rem_cell = ws.cell(row=r, column=remarks_col)
        rem_cell.alignment = right_alignment

        if rem_cell.value not in ["", "-"]:
            rem_cell.fill = light_red_fill
            rem_cell.font = Font(name="Arial", size=12, color="FF0000", bold=False)
        else:
            rem_cell.font = black_font

        # Number format 0.0 for leave columns & PH
        for c_idx in leave_cols + [ph_col]:
            ws.cell(row=r, column=c_idx).number_format = "0.0"

        sn += 1

    # ---------- Totals row ----------
    total_row = start_row + days_in_month + 2
    ws[f"A{total_row}"].value = "Total"
    ws[f"A{total_row}"].font = Font(name="Arial", size=12, bold=False, color="000000")
    ws[f"A{total_row}"].alignment = center_alignment
    ws[f"A{total_row}"].border = thin_border
    ws[f"B{total_row}"].border = thin_border

    col_keys = ["At Work", "Public Holiday", "Sick Leave", "Childcare Leave", "Annual Leave"]
    if ns_leave_present:
        col_keys.append("National Service Leave")

    for offset, key in enumerate(col_keys):
        c_idx = 3 + offset
        val = totals[key]
        cell = ws.cell(row=total_row, column=c_idx, value=("-" if val == 0.0 else val))
        cell.font = Font(name="Arial", size=12, bold=False, color="000000")
        cell.alignment = right_alignment
        cell.border = thin_border
        cell.number_format = "0.0"

    # Border the final remarks column too
    ws[f"{_col_letter(remarks_col)}{total_row}"].border = thin_border

    # ---------- Signature block ----------
    current_date = datetime.now().strftime("%d - %b - %Y")

    # Officer rows
    ws.merge_cells(f"B{total_row + 2}:D{total_row + 2}")
    ws.merge_cells(f"B{total_row + 3}:D{total_row + 3}")
    ws.merge_cells(f"B{total_row + 4}:D{total_row + 4}")

    # Reporting officer rows
    ws.merge_cells(f"B{total_row + 6}:D{total_row + 6}")
    ws.merge_cells(f"B{total_row + 7}:D{total_row + 7}")
    ws.merge_cells(f"B{total_row + 8}:D{total_row + 8}")

    ws[f"A{total_row + 2}"], ws[f"B{total_row + 2}"] = "Officer", name
    ws[f"A{total_row + 3}"], ws[f"B{total_row + 3}"] = "Signature", name
    ws[f"A{total_row + 4}"], ws[f"B{total_row + 4}"] = "Date", current_date

    ws[f"A{total_row + 6}"], ws[f"B{total_row + 6}"] = "Reporting Officer", reporting_officer
    ws[f"A{total_row + 7}"], ws[f"B{total_row + 7}"] = "Signature", ""
    ws[f"A{total_row + 8}"], ws[f"B{total_row + 8}"] = "Date", ""

    ws[f"B{total_row + 4}"].number_format = "DD - MMM - YYYY"
    ws[f"B{total_row + 8}"].number_format = "DD - MMM - YYYY"

    # Borders for A-D in signature area
    for r in range(total_row + 2, total_row + 9):
        for c in ("A", "B", "C", "D"):
            ws[f"{c}{r}"].border = thin_border
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="bottom")
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="bottom")

    # Global font: Arial 12 (don’t overwrite header row or remarks coloring)
    arial = Font(name="Arial", size=12)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if r == 10:
                continue
            # keep Remarks row text coloring
            if (c == remarks_col) and (start_row <= r < start_row + days_in_month):
                continue
            ws.cell(row=r, column=c).font = arial

    # Keep remarks column comfortably wide
    rem_letter = _col_letter(remarks_col)
    ws.column_dimensions[rem_letter].width = 22
    if ns_leave_present:
        ws.column_dimensions["H"].width = 22  # NS column

    # Save
    wb.save(str(out_path))
    return str(out_path)


# Convenience wrapper used by the CLI engine
def generate_cli(
    profile: Dict,
    month_int: int,
    year: int,
    leave_details: List[Sequence],
    out_dir: str | Path = "generated_timesheets",
    public_holidays: Optional[Dict[str, str]] = None,
    remarks: Optional[Dict[str, str]] = None,
) -> str:
    """
    Thin wrapper so your CLI can call a single entrypoint.
    """
    return generate_govtech_timesheet(
        profile=profile,
        month=month_int,
        year=year,
        leave_details=leave_details,
        out_dir=out_dir,
        public_holidays=public_holidays,
        remarks=remarks,
    )
